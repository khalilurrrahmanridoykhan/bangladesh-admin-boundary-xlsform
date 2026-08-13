from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from threading import RLock

from openpyxl import load_workbook

from .models import Area, CrosswalkEntry


PREFIX_LEVEL = {"div": "division", "dis": "district", "upa": "upazila", "uni": "union"}


def bare_code(value: str | None) -> str | None:
    if not value:
        return None
    return value.split("_", 1)[1] if "_" in value else value


def parse_ring(value: str | None) -> list[tuple[float, float]]:
    if not value:
        return []
    points = []
    for vertex in value.split(";"):
        parts = vertex.strip().split()
        if len(parts) >= 2:
            points.append((float(parts[1]), float(parts[0])))  # longitude, latitude
    return points


def point_in_ring(longitude: float, latitude: float, ring: list[tuple[float, float]]) -> bool:
    """Ray-casting point-in-polygon test for the simple Union rings in XLSForm."""
    inside = False
    j = len(ring) - 1
    for i, (xi, yi) in enumerate(ring):
        xj, yj = ring[j]
        if ((yi > latitude) != (yj > latitude)) and (
            longitude < (xj - xi) * (latitude - yi) / ((yj - yi) or 1e-15) + xi
        ):
            inside = not inside
        j = i
    return inside


@dataclass(frozen=True)
class Boundary:
    area_id: str
    ring: list[tuple[float, float]]
    bbox: tuple[float, float, float, float]


class GeoStore:
    def __init__(self, workbook: Path, crosswalk_path: Path):
        self.workbook = workbook
        self.crosswalk_path = crosswalk_path
        self.areas: dict[str, Area] = {}
        self.children: dict[tuple[str, str | None], list[str]] = {}
        self.boundaries: list[Boundary] = []
        self.crosswalk: dict[str, CrosswalkEntry] = {}
        self._lock = RLock()
        self._load_areas()
        self._load_crosswalk()

    def _load_areas(self) -> None:
        ws = load_workbook(self.workbook, read_only=True, data_only=True)["choices"]
        rows = ws.iter_rows(values_only=True)
        headers = [str(value) for value in next(rows)]
        for values in rows:
            row = dict(zip(headers, values))
            prefix = str(row["list_name"]).rstrip("_")
            level = PREFIX_LEVEL[prefix]
            area_id = str(row["name"])
            geo_code = bare_code(area_id)
            parent_field = {"district": "div_filter", "upazila": "dis_filter", "union": "upa_filter"}.get(level)
            parent = str(row.get(parent_field)) if parent_field and row.get(parent_field) else None
            ring = parse_ring(row.get("geometry"))
            area = Area(
                area_id=area_id,
                geo_code=geo_code,
                level=level,
                name=str(row["label"]),
                latitude=float(row["latitude"]),
                longitude=float(row["longitude"]),
                parent_area_id=parent,
                has_boundary=bool(ring),
            )
            self.areas[area_id] = area
            self.children.setdefault((level, parent), []).append(area_id)
            if ring:
                xs, ys = zip(*ring)
                self.boundaries.append(Boundary(area_id, ring, (min(xs), min(ys), max(xs), max(ys))))
        for codes in self.children.values():
            codes.sort(key=lambda code: self.areas[code].name.casefold())

    def _load_crosswalk(self) -> None:
        if self.crosswalk_path.exists():
            raw = json.loads(self.crosswalk_path.read_text())
            self.crosswalk = {item["area_id"]: CrosswalkEntry(**item) for item in raw}

    def list_areas(self, level: str, parent: str | None = None) -> list[Area]:
        return [self.areas[code] for code in self.children.get((level, parent), [])]

    def lineage(self, area_id: str) -> list[Area]:
        result = []
        current = self.areas.get(area_id)
        while current:
            result.append(current)
            current = self.areas.get(current.parent_area_id or "")
        return list(reversed(result))

    def locate(self, latitude: float, longitude: float) -> list[Area]:
        for boundary in self.boundaries:
            minx, miny, maxx, maxy = boundary.bbox
            if minx <= longitude <= maxx and miny <= latitude <= maxy:
                if point_in_ring(longitude, latitude, boundary.ring):
                    return self.lineage(boundary.area_id)
        return []

    def save_crosswalk(self, entry: CrosswalkEntry) -> CrosswalkEntry:
        if entry.area_id not in self.areas:
            raise KeyError(entry.area_id)
        with self._lock:
            self.crosswalk[entry.area_id] = entry
            self.crosswalk_path.parent.mkdir(parents=True, exist_ok=True)
            payload = [item.model_dump() for item in sorted(self.crosswalk.values(), key=lambda x: x.area_id)]
            self.crosswalk_path.write_text(json.dumps(payload, indent=2) + "\n")
        return entry
